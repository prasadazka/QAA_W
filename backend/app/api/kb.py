import io
import csv
import re
import uuid
import logging
from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from pydantic import BaseModel
from app.core.database import get_db
from app.services.embeddings import get_embeddings, get_embedding

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/kb", tags=["knowledge-base"])


# ── Schemas ──────────────────────────────────────────────────

class EntryCreate(BaseModel):
    category_id: str
    question_en: str
    question_ar: str
    answer_en: str
    answer_ar: str
    keywords_en: list[str] = []
    keywords_ar: list[str] = []
    channel: str = "whatsapp_registration"


class EntryUpdate(BaseModel):
    question_en: str | None = None
    question_ar: str | None = None
    answer_en: str | None = None
    answer_ar: str | None = None
    keywords_en: list[str] | None = None
    keywords_ar: list[str] | None = None


# ── Categories ───────────────────────────────────────────────

@router.get("/categories")
def list_categories(channel: str = Query("whatsapp_registration")):
    """List all KB categories."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(
            """SELECT id, name_en, name_ar, description_en, description_ar,
                      channel, sort_order, is_active
               FROM kb_categories WHERE channel = %s ORDER BY sort_order""",
            (channel,),
        )
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


# ── Entries CRUD ─────────────────────────────────────────────

@router.get("/entries")
def list_entries(
    channel: str = Query("whatsapp_registration"),
    category_id: str = Query(None),
    active_only: bool = Query(False),
):
    """List all KB entries with embedding status."""
    with get_db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT e.id, e.question_en, e.question_ar, e.answer_en, e.answer_ar,
                   e.keywords_en, e.keywords_ar, e.channel, e.is_active, e.hit_count,
                   e.source, e.version, e.created_at, e.updated_at,
                   c.name_en AS category_name,
                   (e.embedding IS NOT NULL) AS has_embedding
            FROM kb_entries e
            LEFT JOIN kb_categories c ON e.category_id = c.id
            WHERE e.channel = %s
        """
        params = [channel]

        if category_id:
            sql += " AND e.category_id = %s"
            params.append(category_id)
        if active_only:
            sql += " AND e.is_active = TRUE"

        sql += " ORDER BY c.sort_order, e.created_at"
        cur.execute(sql, params)
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


@router.get("/entries/{entry_id}")
def get_entry(entry_id: str):
    """Get a single KB entry with full details."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(
            """SELECT e.id, e.category_id, e.question_en, e.question_ar,
                      e.answer_en, e.answer_ar, e.keywords_en, e.keywords_ar,
                      e.channel, e.is_active, e.hit_count, e.source, e.version,
                      e.created_at, e.updated_at,
                      c.name_en AS category_name,
                      (e.embedding IS NOT NULL) AS has_embedding
               FROM kb_entries e
               LEFT JOIN kb_categories c ON e.category_id = c.id
               WHERE e.id = %s""",
            (entry_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Entry not found")
        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


@router.post("/entries", status_code=201)
async def create_entry(data: EntryCreate):
    """Create a new KB entry and generate its embedding."""
    entry_id = uuid.uuid4()
    embed_text = f"{data.question_en} {data.question_ar} {data.answer_en[:200]}"

    try:
        embedding = await get_embedding(embed_text)
    except Exception as e:
        logger.warning(f"Embedding generation failed: {e}")
        embedding = None

    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO kb_entries
               (id, category_id, question_en, question_ar, answer_en, answer_ar,
                keywords_en, keywords_ar, channel, source, embedding)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'manual_entry', %s)""",
            (entry_id, data.category_id, data.question_en, data.question_ar,
             data.answer_en, data.answer_ar, data.keywords_en, data.keywords_ar,
             data.channel,
             str(embedding) if embedding else None),
        )

    return {"id": str(entry_id), "has_embedding": embedding is not None}


@router.put("/entries/{entry_id}")
async def update_entry(entry_id: str, data: EntryUpdate):
    """Update a KB entry and regenerate its embedding."""
    with get_db() as conn:
        cur = conn.cursor()

        # Get current entry
        cur.execute("SELECT question_en, question_ar, answer_en FROM kb_entries WHERE id = %s", (entry_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Entry not found")

        q_en = data.question_en or row[0]
        q_ar = data.question_ar or row[1]
        a_en = data.answer_en or row[2]

        # Build SET clause dynamically
        updates = []
        params = []
        if data.question_en is not None:
            updates.append("question_en = %s")
            params.append(data.question_en)
        if data.question_ar is not None:
            updates.append("question_ar = %s")
            params.append(data.question_ar)
        if data.answer_en is not None:
            updates.append("answer_en = %s")
            params.append(data.answer_en)
        if data.answer_ar is not None:
            updates.append("answer_ar = %s")
            params.append(data.answer_ar)
        if data.keywords_en is not None:
            updates.append("keywords_en = %s")
            params.append(data.keywords_en)
        if data.keywords_ar is not None:
            updates.append("keywords_ar = %s")
            params.append(data.keywords_ar)

        if not updates:
            raise HTTPException(400, "No fields to update")

        # Regenerate embedding
        try:
            embedding = await get_embedding(f"{q_en} {q_ar} {a_en[:200]}")
            updates.append("embedding = %s")
            params.append(str(embedding))
        except Exception as e:
            logger.warning(f"Embedding regeneration failed: {e}")

        updates.append("version = version + 1")
        updates.append("updated_at = NOW()")
        params.append(entry_id)

        cur.execute(f"UPDATE kb_entries SET {', '.join(updates)} WHERE id = %s", params)

    return {"status": "updated", "id": entry_id}


@router.delete("/entries/{entry_id}")
def delete_entry(entry_id: str):
    """Delete a KB entry."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM kb_entries WHERE id = %s RETURNING id", (entry_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Entry not found")
    return {"status": "deleted", "id": entry_id}


@router.patch("/entries/{entry_id}/toggle")
def toggle_entry(entry_id: str):
    """Toggle is_active on a KB entry."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE kb_entries SET is_active = NOT is_active, updated_at = NOW() WHERE id = %s RETURNING is_active",
            (entry_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Entry not found")
    return {"id": entry_id, "is_active": row[0]}


# ── Embeddings ───────────────────────────────────────────────

@router.post("/entries/{entry_id}/embed")
async def regenerate_embedding(entry_id: str):
    """Regenerate embedding for a single entry."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT question_en, question_ar, answer_en FROM kb_entries WHERE id = %s", (entry_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Entry not found")

        embed_text = f"{row[0]} {row[1]} {row[2][:200]}"
        embedding = await get_embedding(embed_text)

        cur.execute("UPDATE kb_entries SET embedding = %s WHERE id = %s", (str(embedding), entry_id))

    return {"status": "embedded", "id": entry_id, "dimensions": len(embedding)}


@router.post("/embed-all")
async def embed_all_entries(channel: str = Query("whatsapp_registration")):
    """Regenerate embeddings for all entries missing them."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT id, question_en, question_ar, answer_en FROM kb_entries WHERE embedding IS NULL AND channel = %s",
            (channel,),
        )
        rows = cur.fetchall()

        if not rows:
            return {"status": "ok", "message": "All entries already have embeddings", "count": 0}

        texts = [f"{r[1]} {r[2]} {r[3][:200]}" for r in rows]
        embeddings = await get_embeddings(texts)

        for (entry_id, _, _, _), emb in zip(rows, embeddings):
            cur.execute("UPDATE kb_entries SET embedding = %s WHERE id = %s", (str(emb), entry_id))

    return {"status": "embedded", "count": len(rows)}


# ── File Upload ──────────────────────────────────────────────

SUPPORTED_EXTENSIONS = (".csv", ".xlsx", ".xls", ".pdf", ".txt", ".md", ".docx", ".doc")


def _extract_text_from_pdf(content: bytes) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_text_from_docx(content: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _extract_rows_from_excel(content: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, excel_row)))
    return rows


def _chunk_text(text: str, chunk_size: int = 500) -> list[str]:
    """Split text into chunks by paragraphs/sections, max chunk_size chars."""
    # Split on double newlines (paragraphs) or markdown headers
    sections = re.split(r'\n{2,}|(?=^#{1,3}\s)', text, flags=re.MULTILINE)
    chunks = []
    current = ""

    for section in sections:
        section = section.strip()
        if not section:
            continue
        if len(current) + len(section) > chunk_size and current:
            chunks.append(current.strip())
            current = section
        else:
            current = f"{current}\n\n{section}" if current else section

    if current.strip():
        chunks.append(current.strip())

    return [c for c in chunks if len(c) > 20]  # Skip tiny fragments


def _get_or_create_category(cur, doc_name: str, channel: str) -> str:
    """Auto-create a category from document name if it doesn't exist."""
    # Clean filename to use as category name
    name = re.sub(r'\.[^.]+$', '', doc_name)  # remove extension
    name = re.sub(r'[_\-]+', ' ', name).strip().title()
    if not name:
        name = "General"

    # Check if category already exists
    cur.execute(
        "SELECT id FROM kb_categories WHERE lower(name_en) = lower(%s) AND channel = %s",
        (name, channel),
    )
    row = cur.fetchone()
    if row:
        return row[0]

    # Create new category
    cat_id = uuid.uuid4()
    cur.execute(
        """INSERT INTO kb_categories (id, name_en, name_ar, channel, is_active)
           VALUES (%s, %s, %s, %s, TRUE)""",
        (cat_id, name, name, channel),
    )
    return cat_id


@router.post("/upload")
async def upload_kb_file(
    file: UploadFile = File(...),
    channel: str = Query("whatsapp_registration"),
):
    """Upload a file to create KB entries with auto-generated embeddings.

    Just upload any supported file — the system handles chunking, categorization, and embedding.

    Supported: .csv, .xlsx, .xls, .pdf, .txt, .md, .docx, .doc
    """
    filename = file.filename.lower()
    if not filename.endswith(SUPPORTED_EXTENSIONS):
        raise HTTPException(400, f"Unsupported file type. Supported: {', '.join(SUPPORTED_EXTENSIONS)}")

    content = await file.read()
    if not content:
        raise HTTPException(400, "File is empty")

    # ── Structured files (CSV/Excel) ──
    if filename.endswith((".csv", ".xlsx", ".xls")):
        return await _process_structured_file(filename, content, channel, file.filename)

    # ── Unstructured files (PDF/TXT/MD/DOCX) ──
    return await _process_document_file(filename, content, channel, file.filename)


async def _process_structured_file(filename: str, content: bytes, channel: str, original_name: str):
    """Process CSV/Excel with Q&A columns."""
    if filename.endswith(".csv"):
        rows = _extract_rows_from_csv(content)
    else:
        rows = _extract_rows_from_excel(content)

    if not rows:
        raise HTTPException(400, "File has no data rows")

    required = {"question_en", "question_ar", "answer_en", "answer_ar"}
    missing = required - set(rows[0].keys())
    if missing:
        raise HTTPException(400, f"Missing required columns: {missing}")

    texts = [f"{r['question_en']} {r['question_ar']} {str(r['answer_en'])[:200]}" for r in rows]

    try:
        embeddings = await get_embeddings(texts)
    except Exception as e:
        logger.warning(f"Batch embedding failed: {e}")
        embeddings = [None] * len(rows)

    created = 0
    with get_db() as conn:
        cur = conn.cursor()
        category_id = _get_or_create_category(cur, original_name, channel)

        for row, emb in zip(rows, embeddings):
            entry_id = uuid.uuid4()
            kw_en = [k.strip() for k in str(row.get("keywords_en", "")).split(",") if k.strip()]
            kw_ar = [k.strip() for k in str(row.get("keywords_ar", "")).split(",") if k.strip()]

            cur.execute(
                """INSERT INTO kb_entries
                   (id, category_id, question_en, question_ar, answer_en, answer_ar,
                    keywords_en, keywords_ar, channel, source, embedding)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'file_upload', %s)""",
                (entry_id, category_id,
                 row["question_en"], row["question_ar"],
                 row["answer_en"], row["answer_ar"],
                 kw_en, kw_ar, channel,
                 str(emb) if emb else None),
            )
            created += 1

    return {
        "status": "uploaded",
        "file": original_name,
        "type": "structured",
        "entries_created": created,
        "embeddings_generated": sum(1 for e in embeddings if e is not None),
    }


async def _process_document_file(filename: str, content: bytes, channel: str, original_name: str):
    """Process PDF/TXT/MD/DOCX by chunking and auto-embedding."""
    if filename.endswith(".pdf"):
        text = _extract_text_from_pdf(content)
    elif filename.endswith((".docx", ".doc")):
        text = _extract_text_from_docx(content)
    else:  # .txt, .md
        text = content.decode("utf-8-sig")

    if not text.strip():
        raise HTTPException(400, "Could not extract text from file")

    chunks = _chunk_text(text)
    if not chunks:
        raise HTTPException(400, "No meaningful content found after parsing")

    try:
        embeddings = await get_embeddings(chunks)
    except Exception as e:
        logger.warning(f"Batch embedding failed: {e}")
        embeddings = [None] * len(chunks)

    created = 0
    with get_db() as conn:
        cur = conn.cursor()
        category_id = _get_or_create_category(cur, original_name, channel)

        for i, (chunk, emb) in enumerate(zip(chunks, embeddings)):
            entry_id = uuid.uuid4()
            lines = chunk.split("\n", 1)
            title = lines[0].strip().lstrip("#").strip()[:200]
            body = lines[1].strip() if len(lines) > 1 else chunk

            keywords = [w.lower() for w in re.findall(r'\b[a-zA-Z]{3,}\b', title)][:8]

            cur.execute(
                """INSERT INTO kb_entries
                   (id, category_id, question_en, question_ar, answer_en, answer_ar,
                    keywords_en, keywords_ar, channel, source, source_url, embedding)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'file_upload', %s, %s)""",
                (entry_id, category_id,
                 title, title,
                 body, body,
                 keywords, [],
                 channel, original_name,
                 str(emb) if emb else None),
            )
            created += 1

    return {
        "status": "uploaded",
        "file": original_name,
        "type": "document",
        "chunks": len(chunks),
        "entries_created": created,
        "embeddings_generated": sum(1 for e in embeddings if e is not None),
    }


# ── Documents (uploaded files) ────────────────────────────────

@router.get("/documents")
def list_documents(channel: str = Query("whatsapp_registration")):
    """List all uploaded documents with their entry counts."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT source_url, count(*) AS entries,
                   count(*) FILTER (WHERE embedding IS NOT NULL) AS embedded,
                   min(created_at) AS uploaded_at
            FROM kb_entries
            WHERE channel = %s AND source_url IS NOT NULL
            GROUP BY source_url
            ORDER BY min(created_at) DESC
        """, (channel,))
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


@router.delete("/documents/{doc_name}")
def delete_document(doc_name: str, channel: str = Query("whatsapp_registration")):
    """Delete all KB entries created from a specific uploaded file."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM kb_entries WHERE source_url = %s AND channel = %s RETURNING id",
            (doc_name, channel),
        )
        deleted = len(cur.fetchall())
        if deleted == 0:
            raise HTTPException(404, "No entries found for this document")
    return {"status": "deleted", "document": doc_name, "entries_deleted": deleted}


# ── Stats ────────────────────────────────────────────────────

@router.get("/stats")
def kb_stats(channel: str = Query("whatsapp_registration")):
    """Get KB statistics."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT
                count(*) AS total_entries,
                count(*) FILTER (WHERE is_active) AS active_entries,
                count(*) FILTER (WHERE NOT is_active) AS inactive_entries,
                count(*) FILTER (WHERE embedding IS NOT NULL) AS with_embeddings,
                count(*) FILTER (WHERE embedding IS NULL) AS without_embeddings,
                coalesce(sum(hit_count), 0) AS total_hits
            FROM kb_entries WHERE channel = %s
        """, (channel,))
        row = cur.fetchone()
        cols = [d[0] for d in cur.description]
        stats = dict(zip(cols, row))

        cur.execute("SELECT count(*) FROM kb_categories WHERE channel = %s", (channel,))
        stats["total_categories"] = cur.fetchone()[0]

        return stats
